from io import BytesIO
from django.http import HttpResponse
from django.views import View
from openpyxl import load_workbook
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, PageBreak
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt

@method_decorator(csrf_exempt, name='dispatch')
class ExcelToPDFView(View):
    def post(self, request):
        excel_file = request.FILES.get('excel_file') or request.FILES.get('file')
        if not excel_file:
            return HttpResponse("No Excel file uploaded", status=400)

        wb = load_workbook(excel_file, data_only=True)
        sheet = wb.active

        # Clean Data
        raw_data = [
            [str(cell).strip() if cell is not None else '' for cell in row]
            for row in sheet.iter_rows(values_only=True)
        ]
        data_no_empty_rows = [row for row in raw_data if any(cell.strip() != '' for cell in row)]
        transposed = list(zip(*data_no_empty_rows))
        cleaned_transposed = [col for col in transposed if any(cell.strip() != '' for cell in col)]
        cleaned_data = [list(row) for row in zip(*cleaned_transposed)]

        if not cleaned_data:
            return HttpResponse("Uploaded Excel file has no usable content.", status=400)

        # Generate PDF
        pdf_buffer = BytesIO()
        doc = SimpleDocTemplate(pdf_buffer, pagesize=landscape(A4),
                                leftMargin=20, rightMargin=20, topMargin=30, bottomMargin=20)

        elements = []
        max_cols_per_page = 10
        total_cols = len(cleaned_data[0])

        for col_start in range(0, total_cols, max_cols_per_page):
            col_end = min(col_start + max_cols_per_page, total_cols)
            sliced_data = [row[col_start:col_end] for row in cleaned_data]

            table = Table(sliced_data, repeatRows=1)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.black),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
            ]))

            elements.append(table)
            elements.append(PageBreak())

        doc.build(elements)

        # Return response
        pdf_buffer.seek(0)
        response = HttpResponse(pdf_buffer, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="excel_cleaned.pdf"'
        return response
